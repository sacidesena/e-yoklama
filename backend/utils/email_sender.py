import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from email.header import Header
from datetime import datetime
import os
import io
import re
import traceback
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

SMTP_SERVER = os.getenv("SMTP_SERVER", "smtp.gmail.com")
SMTP_PORT = int(os.getenv("SMTP_PORT", 587))
SMTP_USER = os.getenv("SENDER_EMAIL")
SMTP_PASSWORD = os.getenv("SENDER_PASSWORD")


def create_excel(katilan_ogrenciler, ders_adi, sinif_adi, tarih, baslangic, bitis):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Yoklama Listesi"

    header_fill = PatternFill(start_color="1a3a6e", end_color="1a3a6e", fill_type="solid")
    info_fill = PatternFill(start_color="e8f0fe", end_color="e8f0fe", fill_type="solid")
    row_fill = PatternFill(start_color="f8fafc", end_color="f8fafc", fill_type="solid")

    ws.merge_cells("A1:D1")
    ws["A1"] = "BAİBÜ E-Yoklama Sistemi"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = header_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    info = [
        ("Ders:", ders_adi),
        ("Sınıf:", sinif_adi),
        ("Tarih:", tarih.strftime("%d.%m.%Y")),
        ("Saat:", f"{baslangic} - {bitis}"),
        ("Toplam Katılan:", str(len(katilan_ogrenciler))),
    ]

    for i, (label, value) in enumerate(info, start=2):
        ws[f"A{i}"] = label
        ws[f"B{i}"] = value
        ws[f"A{i}"].font = Font(bold=True)
        ws[f"A{i}"].fill = info_fill
        ws[f"B{i}"].fill = info_fill

    bos_satir = len(info) + 2 + 1
    headers = ["No", "Ad Soyad", "Öğrenci No", "Yoklama Zamanı"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=bos_satir, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[bos_satir].height = 20

    for i, ogrenci in enumerate(katilan_ogrenciler, start=1):
        row = bos_satir + i
        ws.cell(row=row, column=1, value=i).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=2, value=ogrenci['ad'])
        ws.cell(row=row, column=3, value=ogrenci['ogrenci_no']).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=4, value=ogrenci['zaman']).alignment = Alignment(horizontal="center")
        if i % 2 == 0:
            for col in range(1, 5):
                ws.cell(row=row, column=col).fill = row_fill

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 18

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def send_yoklama_list_to_teacher(
    hoca_mail: str,
    hoca_adi: str,
    ders_adi: str,
    ders_kodu: str,
    sinif_adi: str,
    tarih: datetime,
    baslangic: str,
    bitis: str,
    katilan_ogrenciler: list,
    katilmayan_ogrenciler: list
):
    if not SMTP_USER or not SMTP_PASSWORD:
        print("⚠️ Email ayarları yapılmamış, mail gönderilemiyor")
        return False

    toplam = len(katilan_ogrenciler) + len(katilmayan_ogrenciler)
    subject = f"Yoklama Listesi: {ders_adi} - {tarih.strftime('%d.%m.%Y')}"

    html_content = f"""
    <html>
    <head>
        <style>
            body {{ font-family: Arial, sans-serif; background: #f0f4f8; margin: 0; padding: 20px; }}
            .card {{ background: white; border-radius: 12px; overflow: hidden; max-width: 600px; margin: 0 auto; box-shadow: 0 4px 16px rgba(0,0,0,0.1); }}
            .header {{ background: linear-gradient(135deg, #1a3a6e, #2d5a9e); color: white; padding: 24px; text-align: center; }}
            .header h2 {{ margin: 0; font-size: 20px; }}
            .header p {{ margin: 6px 0 0; opacity: 0.8; font-size: 13px; }}
            .info-box {{ background: #f0f4ff; border-left: 4px solid #1a3a6e; margin: 20px; padding: 16px; border-radius: 8px; }}
            .info-box p {{ margin: 6px 0; font-size: 14px; color: #333; }}
            .info-box strong {{ color: #1a3a6e; }}
            .stats {{ display: flex; gap: 12px; margin: 0 20px 20px; }}
            .stat {{ flex: 1; border-radius: 8px; padding: 12px; text-align: center; }}
            .stat-green {{ background: #e8f5e9; }}
            .stat-red {{ background: #fce8e6; }}
            .stat-value-green {{ font-size: 24px; font-weight: bold; color: #2e7d32; }}
            .stat-value-red {{ font-size: 24px; font-weight: bold; color: #c62828; }}
            .stat-label {{ font-size: 12px; color: #666; margin-top: 4px; }}
            table {{ width: calc(100% - 40px); margin: 0 20px 20px; border-collapse: collapse; font-size: 13px; }}
            th {{ background: #1a3a6e; color: white; padding: 10px; text-align: left; }}
            td {{ padding: 9px 10px; border-bottom: 1px solid #eee; }}
            tr:nth-child(even) td {{ background: #f8fafc; }}
            .excel-note {{ background: #fff8e1; border: 1px solid #ffe082; border-radius: 8px; margin: 0 20px 20px; padding: 12px; font-size: 13px; color: #666; }}
            .footer {{ text-align: center; padding: 16px; color: #aaa; font-size: 11px; border-top: 1px solid #eee; }}
        </style>
    </head>
    <body>
        <div class="card">
            <div class="header">
                <h2>📋 Yoklama Listesi</h2>
                <p>Bolu Abant İzzet Baysal Üniversitesi</p>
            </div>

            <div class="info-box">
                <p><strong>👨‍🏫 Sayın {hoca_adi},</strong></p>
                <p><strong>🎓 Ders:</strong> {ders_adi} ({ders_kodu})</p>
                <p><strong>🏫 Sınıf:</strong> {sinif_adi}</p>
                <p><strong>📅 Tarih:</strong> {tarih.strftime('%d.%m.%Y %A')}</p>
                <p><strong>🕐 Saat:</strong> {baslangic} - {bitis}</p>
            </div>

            <div class="stats">
                <div class="stat stat-green">
                    <div class="stat-value-green">{len(katilan_ogrenciler)}</div>
                    <div class="stat-label">Katılan</div>
                </div>
            </div>

            <div class="excel-note">
                📎 Detaylı yoklama listesi Excel dosyası olarak ekte gönderilmiştir.
            </div>

            <table>
                <tr>
                    <th>No</th>
                    <th>Ad Soyad</th>
                    <th>Öğrenci No</th>
                    <th>Saat</th>
                </tr>
    """

    for i, ogrenci in enumerate(katilan_ogrenciler, 1):
        html_content += f"""
                <tr>
                    <td>{i}</td>
                    <td>{ogrenci['ad']}</td>
                    <td>{ogrenci['ogrenci_no']}</td>
                    <td>{ogrenci['zaman']}</td>
                </tr>
        """

    html_content += """
            </table>
            <div class="footer">
                Bu mail otomatik olarak BAİBÜ E-Yoklama Sistemi tarafından gönderilmiştir.
            </div>
        </div>
    </body>
    </html>
    """

    try:
        msg = MIMEMultipart('mixed')
        msg['Subject'] = Header(subject, 'utf-8')
        msg['From'] = f"BAİBÜ E-Yoklama <{SMTP_USER}>"
        msg['To'] = hoca_mail

        html_part = MIMEText(html_content, 'html', 'utf-8')
        msg.attach(html_part)

        # ✅ Dosya adı sanitize
        safe_ders_adi = re.sub(r'[^a-zA-Z0-9]', '_', ders_adi)
        dosya_adi = f"yoklama_{safe_ders_adi}_{tarih.strftime('%d%m%Y')}.xlsx"

        excel_data = create_excel(katilan_ogrenciler, ders_adi, sinif_adi, tarih, baslangic, bitis)
        excel_attachment = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        excel_attachment.set_payload(excel_data)
        encoders.encode_base64(excel_attachment)
        excel_attachment.add_header('Content-Disposition', 'attachment', filename=dosya_adi)
        excel_attachment.add_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        msg.attach(excel_attachment)

        # ✅ ehlo() ile daha stabil bağlantı
        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            server.ehlo()
            server.starttls()
            server.ehlo()
            server.login(SMTP_USER, SMTP_PASSWORD)
            server.send_message(msg)

        print(f"✅ Mail gönderildi: {hoca_mail}")
        return True

    except Exception as e:
        print(f"❌ Mail gönderme hatası: {str(e)}")
        traceback.print_exc()
        return False